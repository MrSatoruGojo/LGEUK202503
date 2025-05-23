# tests/test_formatter.py

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from src.output.formatter import format_workbook

def test_format_workbook(tmp_path):
    # Create a simple Excel file
    file_path = tmp_path / "test.xlsx"
    df = pd.DataFrame({'A': [1,2], 'B': [3,4]})
    df.to_excel(file_path, index=False)

    # Apply formatting
    format_workbook(str(file_path))

    # Reload and inspect
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    hdr = ws.cell(row=1, column=1)

    # Header should be bold
    assert hdr.font.bold
    # Header fill should be solid grey
    
    assert hdr.fill.fill_type == 'solid'
    assert hdr.fill.fgColor.rgb.upper() in ('D3D3D3', '00D3D3D3')

    # Centered alignment
    assert hdr.alignment.horizontal == 'center'
    assert hdr.alignment.vertical   == 'center'
