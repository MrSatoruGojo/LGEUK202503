# src/io/file_ops.py

"""
I/O operations: reading and writing Excel files, folder management,
and listing closed order sources. Supports .xlsx, .xls, and .xlsb.
"""

import os
import glob
import pandas as pd
import openpyxl
from datetime import datetime

# Attempt to import pyxlsb for .xlsb support
try:
    import pyxlsb
except ImportError:
    pyxlsb = None

import config.config as cfg


def create_unique_folder(base_name: str, path: str = None) -> str:
    """
    Create a folder named YYYYMMDD_base_name under `path` (or cfg.OUTPUT_DIR if None).
    Appends _1, _2, ... if the folder already exists.
    """
    if path is None:
        path = cfg.OUTPUT_DIR

    today = datetime.now().strftime('%Y%m%d')
    folder_name = f"{today}_{base_name}"
    full_path = os.path.join(path, folder_name)

    counter = 1
    while os.path.exists(full_path):
        full_path = os.path.join(path, f"{today}_{base_name}_{counter}")
        counter += 1

    os.makedirs(full_path, exist_ok=True)
    return full_path


def read_excel_file(path, sheet_name=None, **kwargs) -> pd.DataFrame:
    """
    Read a single-sheet Excel file (xls, xlsx, or xlsb) into a DataFrame.
    - Unwraps a one-element list into a str.
    - Uses pyxlsb for .xlsb (listing sheets via wb.sheets, engine='pyxlsb').
    - Uses openpyxl for .xlsx/.xls (wb.sheetnames).
    - If sheet_name is provided and exists, uses it; otherwise defaults to the first.
    - Passes additional kwargs (e.g. skiprows) to pd.read_excel.
    """
    # 1) Unwrap single-element lists/tuples
    if isinstance(path, (list, tuple)):
        path = path[0]

    ext = os.path.splitext(path)[1].lower()

    # 2) Determine available sheet-names
    if ext == '.xlsb':
        if pyxlsb is None:
            raise ImportError("pyxlsb is required to read .xlsb files; install via `pip install pyxlsb`")
        wb = pyxlsb.open_workbook(path)
        sheets = wb.sheets  # list of strings
    else:
        # openpyxl for xlsx/xls
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames

    # 3) Choose the sheet to read
    if sheet_name and sheet_name in sheets:
        target = sheet_name
    else:
        target = sheets[0]

    # 4) Build kwargs for pandas.read_excel
    read_args = {'sheet_name': target, **kwargs}
    if ext == '.xlsb':
        read_args['engine'] = 'pyxlsb'

    # 5) Read and return
    return pd.read_excel(path, **read_args)


def write_excel_file(df: pd.DataFrame, path: str, sheet_name: str = 'Sheet1', index: bool = False):
    """
    Write a DataFrame to Excel, creating parent dirs if needed.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df.to_excel(path, sheet_name=sheet_name, index=index)


def read_data(path: str, sheet_name=None, **kwargs):
    """
    Universal data reader:
      - Excel (.xls/.xlsx/.xlsb) → uses read_excel_file()
      - CSV  (.csv)              → pd.read_csv()
      - JSON (.json)             → pd.read_json()
    """
       # If path is a list (e.g. CLOSED_ORDERS_DIR), return it unchanged
    if isinstance(path, (list, tuple)):
        return path
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xls', '.xlsx', '.xlsb'):
        return read_excel_file(path, sheet_name=sheet_name, **kwargs)
    elif ext == '.csv':
        import pandas as pd
        return pd.read_csv(path, **kwargs)
    elif ext == '.json':
        import pandas as pd
        return pd.read_json(path, **kwargs)
    else:
        raise ValueError(f"Unsupported file extension: {ext}")    

def list_closed_order_files() -> list[str]:
    """
    Return the list of closed-orders files:
    - If cfg.CLOSED_ORDERS_DIR is a list, return it.
    - If it’s a string pointing to a folder, glob for *.xls, *.xlsx, *.xlsb.
    """
    files = cfg.CLOSED_ORDERS_DIR
    if isinstance(files, (list, tuple)):
        return files
    elif isinstance(files, str) and os.path.isdir(files):
        patterns = ['*.xls', '*.xlsx', '*.xlsb']
        out = []
        for p in patterns:
            out.extend(glob.glob(os.path.join(files, p)))
        return sorted(set(out))
    else:
        raise ValueError(f"CLOSED_ORDERS_DIR must be a list of files or a folder path, not {files!r}")


def get_export_path(filename_template: str = None) -> str:
    """
    Build a timestamped export filename under cfg.OUTPUT_DIR.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if filename_template:
        name = filename_template.format(timestamp=timestamp)
    else:
        name = f"export_{timestamp}.xlsx"
    return os.path.join(cfg.OUTPUT_DIR, name)
