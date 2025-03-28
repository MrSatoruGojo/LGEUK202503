#1 Importing everything

import pandas as pd
import glob
import time
from tqdm import tqdm
import logging
import os
import getpass
import re
import openpyxl
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.dataframe import dataframe_to_rows






































# 2 Set up logging for time (LOGGING of progress bar)
logging.basicConfig(level=logging.INFO, filename='process_log.txt', filemode='w', format='%(asctime)s - %(levelname)s - %(message)s')

# 3 FOLDER CREATION FOR PROJECT Function to create a unique folder in the provided path
def create_unique_folder(base_name, path):
    today_date = datetime.today().strftime('%Y%m%d')
    base_folder = f"{today_date} {base_name}"
    folder_path = os.path.join(path, base_folder)
    
    # If folder exists, create a new folder with incremental number
    folder_number = 1
    while os.path.exists(folder_path):
        folder_path = os.path.join(path, f"{today_date} {base_name} ({folder_number})")
        folder_number += 1
    
    # Create the folder
    os.makedirs(folder_path)
    return folder_path

# 4.1 CREATING DICTIONARY (2 defs) Create dictionary of SPMS for lookups
def create_lookup_dict(df, key_col, value_cols):
    """
    Converts a dataframe into a dictionary lookup table.
    Handles duplicate keys by storing lists of values instead of single values
    """
    grouped = df.groupby(key_col)[value_cols].apply(lambda x: x.to_dict(orient='records')).to_dict()
    return grouped

# 4.2 CREATING DICTIONARY
def lookup_value(Promo_No, field_name, lookup_dict, default_value="Not Found"):
    """
    Retrieves the first occurrence of a specific field for a given promotion
    """
    records = lookup_dict.get(Promo_No, [])
    return records[0].get(field_name, default_value) if records else default_value

def lookup_customer(Promo_No, field_name, lookup_dict, customer_name=None, secondary_lookup_dict=None, default_value="Not Found"):
    """
    Retrieves the first occurrence of a specific field for a given promotion.
    First checks the main lookup dictionary (`lookup_dict`).
    If not found, checks the secondary lookup dictionary (`secondary_lookup_dict`).
    """
    if customer_name:
        # Perform lookup in the primary dictionary (SPMS)
        records = lookup_dict.get(Promo_No, [])
        records = [record for record in records if record.get('Bill To Name', '')[:12] == customer_name]

        # If no match is found, try the secondary lookup dictionary (SPMS2)
        if not records and secondary_lookup_dict:
            records = secondary_lookup_dict.get(Promo_No, [])
            records = [record for record in records if record.get('Bill To Name', '')[:12] == customer_name]
    
    else:
        # If no customer_name is passed, use Promo_No for the lookup
        records = lookup_dict.get(Promo_No, [])
    
    return records[0].get(field_name, default_value) if records else default_value

def lookup_customer2(Promo_No, field_name, lookup_dict, product_code=None, secondary_lookup_dict=None, default_value="Not Found"):
    """
    Retrieves the first occurrence of a specific field for a given promotion.
    First checks the main lookup dictionary (`lookup_dict`).
    If not found, checks the secondary lookup dictionary (`secondary_lookup_dict`).
    """
    if product_code:
        # Perform lookup in the primary dictionary (SPMS)
        records = lookup_dict.get(Promo_No, [])
        records = [record for record in records if record.get('Product Code','') == product_code]

        # If no match is found, try the secondary lookup dictionary (SPMS2)
        if not records and secondary_lookup_dict:
            records = secondary_lookup_dict.get(Promo_No, [])
            records = [record for record in records if record.get('Product Code', '') == product_code]
    
    else:
        # If no customer_name is passed, use Promo_No for the lookup
        records = lookup_dict.get(Promo_No, [])
    
    return records[0].get(field_name, default_value) if records else default_value


def lookup_SPGM(Promo_No, Bill_To_Name, lookup_dict, secondary_lookup_dict=None, default_value="Not Found"):
    """
    Retrieves 'Sales PGM NO' based on Promotion No and Bill To Name.
    Checks primary lookup_dict first, then secondary_lookup_dict if not found.
    """
    # Check primary dictionary with matching Bill To Name
    records = lookup_dict.get(Promo_No, [])
    filtered_records = [record for record in records if record.get('Bill To Name', '')[:12] == Bill_To_Name[:12]]

    # If not found in primary dictionary, check secondary
    if not filtered_records and secondary_lookup_dict:
        records = secondary_lookup_dict.get(Promo_No, [])
        filtered_records = [record for record in records if record.get('Bill To Name', '')[:12] == Bill_To_Name]

    return filtered_records[0].get('Sales PGM NO', default_value) if filtered_records else default_value


# 5.1 WEEK CONVERTATION - Function to convert YYYYMMDD to week number
def date_to_week(date_str):
    date = pd.to_datetime(str(date_str), format='%Y%m%d')
    return date.isocalendar()[1]  # This returns the week number

# 5.2 WEEK CONVERTATION - Function to extract year from YYYYMMDD
def date_to_year(date_str):
    date = pd.to_datetime(str(date_str), format='%Y%m%d')
    return date.year

# 6.1 GIVING FILES SOURCES - SOURCE FILES CALL
file_claim = (r'E:\1. SFM\0. Main Files\Projects\2.1. CLAIM THING\DRAFT5.xlsx')
file_SPMS = (r'E:\1. SFM\0. Main Files\Projects\2.1. CLAIM THING\3. DATABASE 3 - SPMS DATA\SPMS.xlsx')
file_PSI_data = (r'E:\1. SFM\0. Main Files\Projects\2.1. CLAIM THING\2.DATABASE 2 - PSI DATA\PSI.xlsx')  # New dataset path
file_newtracker = 3
file_oldtracker = (r'E:\1. SFM\0. Main Files\Projects\2.1. CLAIM THING\4. DATABASE 4 - TRACKER DATA\TV1NF.xlsx')

# 6.2 GIVING FILES SOURCES - Closed Orders Files
closed_orders_files = [
    r'V:\2020 CLOSED ORDERS 01.01.20 - 31.12.20.xlsx',
    r'V:\2021 CLOSED ORDERS 01.01.21 - 31.12.21.xlsx',
    r'V:\2022 CLOSED ORDERS 01.01.22 - 31.12.22.xlsx',
    r'V:\2023 CLOSED ORDERS 01.01.23 - 31.12.23.xlsx',
    r'V:\2024 CLOSED ORDERS 01.01.24 - 31.12.24.xlsx',
    r'V:\2025 CLOSED ORDERS 01.01.25 - 28.02.25.xls.xlsx'
]



# DATASETS CALL (READ)
data_claim = pd.read_excel(file_claim, sheet_name='CLAIM')
data_SPMS = pd.read_excel(file_SPMS, sheet_name='Report 1', skiprows=3)
data_SPMS2 = pd.read_excel(file_SPMS, sheet_name='Report 7', skiprows=2)
claim_AR = pd.read_excel(file_claim, sheet_name='AR')
# Read the PSI dataset into a DataFrame
data_PSI = pd.read_excel(file_PSI_data)
data_oldtracker = pd.read_excel(file_oldtracker, skiprows=1)

# 1.1.2. Convert SPMS into dictionary
spms_mapping = create_lookup_dict(data_SPMS, 'Promotion No', ['Promotion Start YYYYMMDD', 'Promotion End YYYYMMDD', 'Promotion Name', 'Promotion Status Code', 'Cancel Flag', 'Recreate Flag', 'Original Promotion No', 'Sales PGM NO', 'Sales PGM Status', 'Promotion Property', 'Alloc Div Code', 'Apply Month_YYYYMM', 'Bill To Name', 'Claim Line Flag', 'Customer Code', 'Division Code', 'Product Code', 'Expected Qty', 'Dc Operand', 'Expected Cost'])
spms_mapping_secondary = create_lookup_dict(data_SPMS2, 'Promotion No', ['Sales PGM NO','Bill To Name', 'Product Code','Expected Qty', 'Dc Operand','Expected Cost'])
 
data_claim['Promotion number check'] = data_claim['Promotion No'].apply(lambda x: '1' if x in spms_mapping else 'Not found in SPMS')























#print (data_claim)

def promocheck(row, spms_mapping):
    if 'Promotion number check' in row and row['Promotion number check'] == 'Not found in SPMS':
        row['Cancel Flag'] = 0
        row['Recreate Flag'] = 0
        row['Promotion Start Date'] = 0
        row['Promotion End Date'] = 0
        row['Promotion Start Week'] = 0
        row['Promotion End Week'] = 0
        row['Promotion Start Year'] = 0
        row['Promotion End Year'] = 0
        row['Bill To Name SPMS'] = 0
        row['Bill To Name Closed Orders'] = 0
        row['Bill To Name Closed 12'] = 0
        row['Product Code SPMS'] = 0
        row['Sales PGM NO'] = 0
    else:
        row['Cancel Flag'] = '1' if lookup_value(row['Promotion No'], 'Cancel Flag', spms_mapping) == 'Y' else '0'
        row['Recreate Flag'] = '1' if lookup_value(row['Promotion No'], 'Recreate Flag', spms_mapping) == 'Y' else '0'
        # Add promotion start and end date and convert to weeks
        row['Promotion Start Date'] = lookup_value(row['Promotion No'], 'Promotion Start YYYYMMDD', spms_mapping)
        row['Promotion End Date'] = lookup_value(row['Promotion No'], 'Promotion End YYYYMMDD', spms_mapping)

        # Convert the dates to week numbers
        row['Promotion Start Week'] = date_to_week(row['Promotion Start Date'])
        row['Promotion End Week'] = date_to_week(row['Promotion End Date'])

        # Convert the dates to years
        row['Promotion Start Year'] = date_to_year(row['Promotion Start Date'])
        row['Promotion End Year'] = date_to_year(row['Promotion End Date'])
        row['Bill To Name SPMS'] = lookup_customer(row['Promotion No'], 'Bill To Name', spms_mapping, 
                                    customer_name=row['Bill To Name'][:12], secondary_lookup_dict=spms_mapping_secondary)
        row['Bill To Name Closed Orders'] = row['Bill To Name SPMS'][:12]
        row['Product Code SPMS'] = lookup_customer2(row['Promotion No'], 'Product Code', spms_mapping, 
                                    product_code=row['Product Code'], secondary_lookup_dict=spms_mapping_secondary)
        row['Sales PGM NO'] = lookup_SPGM(
    row['Promotion No'],
    row['Bill To Name'][:12],
    spms_mapping,
    secondary_lookup_dict=spms_mapping_secondary
)

        #row['Bill To Name2'] = lookup_value(row['Promotion No'], 'Customer Name', spms_mapping)
    return row

# Check if the 'Product Code' exists in data_claim and SPMS
def check_model_exist(row):
    Promo_No, Product_Code = row['Promotion No'], row['Product Code']  # Ensure column name is 'Product Code'
    spms_models = data_SPMS[data_SPMS['Promotion No'] == Promo_No]['Product Code'].tolist()  # Use 'Product Code' in SPMS
    return '1' if Product_Code in spms_models else '0'

data_claim = data_claim.apply(lambda row: promocheck(row, spms_mapping), axis=1)
data_claim['Model Check'] = data_claim.apply(check_model_exist, axis=1)
#print (data_claim)
#print(data_claim.shape)  # This will show (number of rows, number of columns)
#print(data_claim.head())  # This will show the first 5 rows


save_path = r'E:\1. SFM\0. Main Files\Projects\2.1. CLAIM THING\13. New Logs 20250305'

# Create the unique folder in the user-provided path
folder_path = create_unique_folder('Log', save_path)  # Using the fixed save path

# Paths for output in the newly created folder
file_path_1 = os.path.join(folder_path, '1. CLAIMDATA.xlsx')

data_claim.to_excel(file_path_1, index=False)

print(f"File saved to: {file_path_1}")
print("Stage 1 is finished.")
print("Stage 2 is started")
















































# Placeholder for combined filtered orders
aggregated_orders = []

# Start timing the process
start_time = time.time()

# Ensure short name column exists in data_claim for merging
data_claim['Bill To Name Short'] = data_claim['Bill To Name Closed Orders'].str[:12].str.upper()

# Loop through each file and process it
for file in tqdm(closed_orders_files, desc="Processing Closed Orders Files"):
    try:
        year = file.split(' CLOSED ORDERS')[0].split('\\')[-1]  # Extract YEAR from file name
        closed_orders = pd.read_excel(file, sheet_name=None)

        # Read the only sheet and force a copy to avoid SettingWithCopyWarning
        closed_orders_filtered = closed_orders[list(closed_orders.keys())[0]].copy()

        # Shorten the claim customer names and models for filtering
        claim_customers = data_claim['Bill To Name Closed Orders'].dropna().unique()
        claim_models = set(data_claim['Product Code SPMS'])

        # ✅ DEBUG: Print customer codes and models from claim
        print(f"\n📁 Processing file: {file}")
        print("🔍 Claim Customers (Short):")
        for cust in claim_customers:
            print("-", cust[:12].upper())

        print("\n🔍 Claim Models:")
        print(claim_models)

        # Create a new column for the first 12 characters
        closed_orders_filtered['Bill To Name Short'] = closed_orders_filtered['Bill To Name'].str[:12]

        # ✅ DEBUG: Print the first few rows before filtering
        print("\n📄 FIRST FEW ROWS FROM CURRENT FILE:")
        print(closed_orders_filtered[['Bill To Name Short', 'Model']].head(10))

        # Prepare the set of short customer names for comparison
        claim_customers_short = {cust[:12].upper() for cust in claim_customers}

        # Filter rows that match both customer and model
        closed_orders_filtered = closed_orders_filtered[
            closed_orders_filtered['Bill To Name Short'].apply(lambda x: x.upper() in claim_customers_short) &
            closed_orders_filtered['Model'].isin(claim_models)
        ]

        if closed_orders_filtered.empty:
            print("⚠️ No matching rows after filtering! Check names or models.")
            logging.warning(f"No matching records found in {file}. Skipping file.")
            continue  # Skip this file if no matching records found

        # Save filtered data to Excel for review
        file_path_2 = os.path.join(folder_path, f'2. Closed_orders_filtered_{year}.xlsx')
        closed_orders_filtered.to_excel(file_path_2, index=False)

        print("✅ Matches found. Continuing processing...")

        # Group and sum quantities per customer & model
        closed_orders_filtered['Bill To Name Short'] = closed_orders_filtered['Bill To Name'].str[:12]
        orders_aggregated = closed_orders_filtered.groupby(['Bill To Name Short', 'Model'])['Order Qty'].sum().reset_index()
        orders_aggregated['Year'] = year

        # 🔁 Merge directly into data_claim with a per-year column
        merge_result = data_claim.merge(
            orders_aggregated,
            how='left',
            left_on=['Bill To Name Short', 'Product Code SPMS'],
            right_on=['Bill To Name Short', 'Model']
        )

        merge_result[f'Order Qty {year}'] = merge_result['Order Qty'].fillna(0).astype(int)

        # Clean up temporary columns
        merge_result.drop(columns=['Order Qty', 'Model', 'Year'], inplace=True)

        # Replace data_claim with updated version
        data_claim = merge_result

        # Log processing success
        logging.info(f"Successfully processed file: {file}")

    except Exception as e:
        logging.error(f"Error processing file {file}: {str(e)}")

# Save updated data_claim with added yearly columns
final_output_path = os.path.join(folder_path, '3. data_claim_with_orders.xlsx')
# Identify all columns that match the "Order Qty <year>" pattern
order_qty_columns = [col for col in data_claim.columns if col.startswith('Order Qty ')]

# Create a new column 'Total Closed Orders' by summing across the year columns
data_claim['Total Closed Orders'] = data_claim[order_qty_columns].sum(axis=1)

data_claim.to_excel(final_output_path, index=False)

print(f"\n✅ Final updated data_claim saved to: {final_output_path}")
print("Stage 2 is finished")
print("Stage 3 is starting")




data_claim['Promotion Start Date Formatted'] = data_claim.apply(
    lambda row: f"{str(row['Promotion Start Date'])[2:4]}-(W{row['Promotion Start Week']})"
    if not row['Promotion Start Date'] == 0 or row['Promotion Start Week'] == 'Not Found' 
    else 'Not Found', axis=1)

data_claim['Promotion End Date Formatted'] = data_claim.apply(
    lambda row: f"{str(row['Promotion End Date'])[2:4]}-(W{row['Promotion End Week']})"
    if not row['Promotion End Date'] == 0 or row['Promotion End Week'] =='Not Found'
    else 'Not Found', axis=1)




# Function to generate weeks range between start and end dates with Monday as the start of the week
def generate_weeks_range_monday(start_date, end_date):
    weeks_range = []
    start_week = pd.to_datetime(start_date, format='%Y%m%d')
    end_week = pd.to_datetime(end_date, format='%Y%m%d')

    # Adjust start week to the previous Monday (or itself if already Monday)
    start_week -= pd.DateOffset(days=start_week.weekday())  # Moves to the previous Monday
    
    # Loop to generate weeks in range
    while start_week <= end_week:
        week_start = start_week.strftime('%y-%m-%d')
        week_number = start_week.isocalendar()[1]
        weeks_range.append(f"{week_start}\n(W{week_number})")
        start_week += pd.DateOffset(weeks=1)  # Increment to next week
    
    return weeks_range


# Add the "Week's range" column using Monday as the start of the week
data_claim['Week\'s range'] = data_claim.apply(
    lambda row: ', '.join(generate_weeks_range_monday(row['Promotion Start Date'], row['Promotion End Date'])) 
    if isinstance(row['Promotion Start Date'], (int, float)) and isinstance(row['Promotion End Date'], (int, float)) 
    and row['Promotion Start Date'] != 0 and row['Promotion End Date'] != 0
    else "Not Found", axis=1)

# Add the new columns (SELL-IN, SELL-OUT, INVENTORY) - Placeholder logic
data_claim['SELL-IN'] = 0  # Replace with actual data mapping or calculation
data_claim['SELL-OUT'] = 0  # Replace with actual data mapping or calculation
data_claim['INVENTORY'] = 0  # Replace with actual logic or calculation


# Paths for output in the newly created folder
output_path = os.path.join(folder_path, '4. Data_claim.xlsx')
closed_orders_filtered_output = os.path.join(folder_path, 'closed_orders_filtered.xlsx')

output_path_1 = os.path.join(folder_path, '5. Data_claim_UPD1.xlsx')
data_claim.to_excel(output_path_1, index=False)


























# Check if necessary columns exist in the PSI data
if 'Channel' not in data_PSI.columns or 'Model.Suffix' not in data_PSI.columns:
    print("Error: One or more required columns ('Channel' or 'Model.Suffix') are missing from PSI data.")
else:
    # Filter the PSI data where 'Measure' is 'Sell-Out FCST_KAM [R+F]'
    filtered_data_SO = data_PSI[data_PSI['Measure'] == 'Sell-Out FCST_KAM [R+F]']
    measure_SO = 'Sell-Out FCST_KAM [R+F]'
    print("Filtered PSI data based on 'Sell-Out FCST_KAM [R+F]'")

    # Create a new column in PSI data with the first 12 characters of Channel
    filtered_data_SO['Channel Short'] = filtered_data_SO['Channel'].str[:12]

    # Dynamically select columns starting from column G to the last column (weekly columns)
    weekly_columns_SO = data_PSI.columns[6:]  # Assuming column G is the 7th column (index 6) and we want all columns after it
    print(f"Weekly columns selected: {weekly_columns_SO}")

    # Iterate through the rows in the verification sheet to process each row
    for _, verification_row in data_claim.iterrows():
        # Extract first 12 characters of 'Bill To Name' in verification data
        customer_name_short = verification_row['Bill To Name Short']
        product_code = verification_row['Product Code']
        print(f"Processing Customer Name (Short): {customer_name_short}, Product Code: {product_code}.")

        # Convert 'Week's range' from "25-01-13 (W3)" to "25-01-13\n(W3)" before processing
        weeks_range_SO = [week.replace(" (", "\n(").replace(")", ")") for week in verification_row['Week\'s range'].split(', ')]
        print(f"Weeks selected (after conversion): {weeks_range_SO}")

        # Filter the PSI data based on the first 12 characters of 'Channel' and 'Product Code'
        relevant_psi_data_SO = filtered_data_SO[
            (filtered_data_SO['Channel Short'] == customer_name_short) &
            (filtered_data_SO['Model.Suffix'] == product_code) &
            (filtered_data_SO['Measure'] == measure_SO)
        ]

        # Extract the relevant week columns based on the week's range in the verification sheet
        relevant_columns_SO = [week for week in weeks_range_SO if week in weekly_columns_SO]
        print(f"Relevant columns for summing: {relevant_columns_SO}")

        # If relevant columns are found in the PSI data (i.e., matching weeks), proceed to sum them
        if relevant_columns_SO:
            # Sum the relevant weekly columns in the filtered PSI data
            summed_value_SO = relevant_psi_data_SO[relevant_columns_SO].sum().sum()

            # Update the SELL-OUT column for the relevant rows in the verification sheet with the summed value
            data_claim.loc[
                (data_claim['Bill To Name Short'] == customer_name_short) &
                (data_claim['Product Code'] == product_code),
                'SELL-OUT'
            ] = summed_value_SO
            print(f"SELL-OUT value updated to: {summed_value_SO}")

# Save the updated verification data
file_path4 = os.path.join(folder_path, '6. Sell-out updated.xlsx')
data_claim.to_excel(file_path4, index=False)
print("Stage 4 is finished")
print("Stage 5 is starting")





























###2SELLIN
# Check if necessary columns exist in the PSI data

    # Filter the PSI data where 'Measure' is 'Sell-In FCST_KAM [R+F]'
filtered_data_SI = data_PSI[data_PSI['Measure'] == 'Sell-In FCST_KAM [R+F]']
measure_SI = 'Sell-In FCST_KAM [R+F]'
print("Filtered PSI data based on 'Sell-In FCST_KAM [R+F]' measure.")

    # Create a new column in PSI data with the first 12 characters of Channel
filtered_data_SI['Channel Short'] = filtered_data_SI['Channel'].str[:12]

    # Dynamically select columns starting from column G to the last column (weekly columns)
weekly_columns_SI = data_PSI.columns[6:]  # Assuming column G is the 7th column (index 6) and we want all columns after it
print(f"Weekly columns selected: {weekly_columns_SI}") 

    # Iterate through the rows in the verification sheet to process each row
for _, verification_row_SI in data_claim.iterrows():
        # Extract first 12 characters of 'Bill To Name' in verification data
    customer_name_short = verification_row_SI['Bill To Name Short']
    product_code = verification_row_SI['Product Code']
    print(f"Processing Customer Name (Short): {customer_name_short}, Product Code: {product_code}.")

        # Convert 'Week's range' from "25-01-13 (W3)" to "25-01-13\n(W3)" before processing
    weeks_range_SI = [week.replace(" (", "\n(").replace(")", ")") for week in verification_row_SI['Week\'s range'].split(', ')]
    print(f"Weeks selected (after conversion): {weeks_range_SI}")

        # Filter the PSI data based on the first 12 characters of 'Channel' and 'Product Code'
    relevant_psi_data_SI = filtered_data_SI[
        (filtered_data_SI['Channel Short'] == customer_name_short) &
        (filtered_data_SI['Model.Suffix'] == product_code) &
        (filtered_data_SI['Measure'] == measure_SI)
    ]

        # Extract the relevant week columns based on the week's range in the verification sheet
    relevant_columns_SI = [week for week in weeks_range_SI if week in weekly_columns_SI]
    print(f"Relevant columns for summing: {relevant_columns_SI}")

        # If relevant columns are found in the PSI data (i.e., matching weeks), proceed to sum them
    if relevant_columns_SI:
            # Sum the relevant weekly columns in the filtered PSI data
        summed_value_SI = relevant_psi_data_SI[relevant_columns_SI].sum().sum()

            # Update the SELL-IN column for the relevant rows in the verification sheet with the summed value
        data_claim.loc[
            (data_claim['Bill To Name Short'] == customer_name_short) & 
            (data_claim['Product Code'] == product_code), 
                'SELL-IN'
        ] = summed_value_SI
        print(f"SELL-IN value updated to: {summed_value_SI}")


# Save the updated verification data
file_path5 = os.path.join(folder_path, '7. Sell-IN updated.xlsx')
data_claim.to_excel(file_path5, index=False)
print("Stage 5 is finished")
print("Stage 6 is starting")

































#-3. INVENTORY CALCULATION ##INV

#-3.1 FINDING RELEVANT INVENTORY VALUES
#-3.1.1 Check if necessary columns exist in the PSI data, if not - else

    #-3.1.2 Filter the PSI data where 'Measure' is 'Ch. Inventory_Sellable'
filtered_data_INV = data_PSI[data_PSI['Measure'] == 'Ch. Inventory_Sellable']
measure_INV = 'Ch. Inventory_Sellable'
print("Filtered PSI data based on 'Ch. Inventory_Sellable' measure.")

    # Create a new column in PSI data with the first 12 characters of Channel
filtered_data_INV['Channel Short'] = filtered_data_INV['Channel'].str[:12]

    #-3.1.4 Dynamically select columns starting from column G to the last column (weekly columns)
weekly_columns_INV = data_PSI.columns[6:]  # Assuming column G is the 7th column (index 6) and we want all columns after it
print(f"Weekly columns selected: {weekly_columns_INV}") 

    #-3.2 ITERATING ALL MATCHES 
    # Now let's update the INVENTORY column in the verification sheet based on the PSI data

    #-3.2.1 Iterate through the rows in the verification sheet to process each row
for _, verification_row_INV in data_claim.iterrows():
        #-3.2.2 Get the first 12 characters of 'Bill To Name' and 'Product Code'
    customer_name_short = verification_row_INV['Bill To Name Short']
    product_code = verification_row_INV['Product Code']
    print(f"Processing Customer Name (Short): {customer_name_short}, Product Code: {product_code}.")

        #-3.3 FILTERING AND FORMATTING 
        #-3.3.1 Convert 'Week's range' from "25-01-13 (W3)" to "25-01-13\n(W3)" before processing
    weeks_range_INV = [week.replace(" (", "\n(").replace(")", ")") for week in verification_row_INV['Week\'s range'].split(', ')]
    print(f"Weeks selected (after conversion): {weeks_range_INV}")

        #-3.3.2 Filter the PSI data based on the first 12 characters of 'Channel' and 'Product Code'
    relevant_psi_data_INV = filtered_data_INV[
        (filtered_data_INV['Channel Short'] == customer_name_short) &
        (filtered_data_INV['Model.Suffix'] == product_code) &
        (filtered_data_INV['Measure'] == measure_INV)
    ]

        #-3.3.3 Extract the relevant week columns based on the week's range in the verification sheet
    relevant_columns_INV = [week for week in weeks_range_INV if week in weekly_columns_INV]
    print(f"Relevant columns for summing: {relevant_columns_INV}")

        #-3.4 SUMMING
        #-3.4.1 If relevant columns are found in the PSI data (i.e., matching weeks), proceed to sum them
    if relevant_columns_INV:
            #-3.4.2 Sum the relevant weekly columns in the filtered PSI data
        summed_value_INV = relevant_psi_data_INV[relevant_columns_INV].sum().sum()

            #-3.4.3 Update the INVENTORY column for the relevant rows in the verification sheet with the summed value
        data_claim.loc[
            (data_claim['Bill To Name Short'] == customer_name_short) &
            (data_claim['Product Code'] == product_code), 
                'INVENTORY'
        ] = summed_value_INV

        print(f"INVENTORY value updated to: {summed_value_INV}")

file_path6 = os.path.join(folder_path, '8. Inventory updated.xlsx')
data_claim.to_excel(file_path6, index=False)
print("Stage 6 is finished")
print("Stage 7 is starting")











#-3. TRACKER CALCULATION OLD ##TR

# Ensure 'Customer Short' column in data_oldtracker for 12-character matching
data_oldtracker['Customer Short'] = data_oldtracker['Customer'].astype(str).str[:12]

# Convert 'Claim Volume' to numeric once before processing (ignore strings)
data_oldtracker['Claim Volume'] = pd.to_numeric(data_oldtracker['Claim Volume'], errors='coerce')

# Initialize Tracker column in verification_data with zeros
data_claim['Tracker'] = 0

# Iterate over each row in verification_data to calculate the sum for each line
for idx, verification_row in data_claim.iterrows():
    customer_short = verification_row['Bill To Name Short']  # Already 12 characters
    product_code = verification_row['Product Code SPMS']  # Product code to match

    # Filter data_oldtracker dataset based on matching Customer Short (first 12 characters) and Model
    matched_oldtracker_rows = data_oldtracker[
        (data_oldtracker['Customer Short'] == customer_short) &
        (data_oldtracker['Model'] == product_code)
    ]

    # Keep only numeric values (ignore strings like 'CANCELLED')
    matched_oldtracker_rows = matched_oldtracker_rows[matched_oldtracker_rows['Claim Volume'].notna()]

    # Sum only numeric values
    total_claim_volume = matched_oldtracker_rows['Claim Volume'].sum()

    # Insert summed value into Tracker column in verification_data
    data_claim.at[idx, 'Tracker'] = total_claim_volume

# Print confirmation message
print("Tracker column updated successfully using data_oldtracker!")



#Verification part

data_claim['PSI TOTAL'] = data_claim['SELL-OUT'] + data_claim['INVENTORY'] + data_claim['SELL-IN']
data_claim['CO CHECK'] = data_claim['Total Closed Orders'] - data_claim['Tracker'] - data_claim['Q']    #Closed orders - Tracker - Q
data_claim['PSI CHECK'] = data_claim['SELL-OUT'] - data_claim['Q']
data_claim['ID PSI CHECK'] = data_claim['PSI TOTAL'] - data_claim['Q']

output_path_2 = os.path.join(folder_path, '9. Data_cleaned.xlsx')
#Moving to Excel
data_claim.to_excel(output_path_2, index=False)






















#NEW dataframe for verification of each claim


data_claim_verification = data_claim.copy()


#-2.1 Dropping unneccesarry columns
data_claim_cleaned = data_claim.drop(columns=['Promotion Start Week','Promotion End Week','Promotion Start Year','Promotion End Year','Promotion Start Date Formatted','Promotion End Date Formatted','Week\'s range'])
    
output_path_4 = os.path.join(folder_path, '10. Data_verified.xlsx')
#Moving to Excel
data_claim_cleaned.to_excel(output_path_4, index=False)

output_path = os.path.join(folder_path, '4. Data_claim.xlsx')

















comment_summary = []

# Group by BEBS and Product Code SPMS (model)
for (bebs, model), group in data_claim_verification.groupby(['BEBS', 'Product Code SPMS']):
    
    ok_rows = []
    not_ok_rows = []
    ok_total = 0
    not_ok_total = 0
    comments = []

    for idx, row in group.iterrows():
        # Normalize all fields to strings for safe comparison
        cancel_flag = str(row['Cancel Flag']).strip()
        promo_check = str(row['Promotion number check']).strip()
        model_check = str(row['Model Check']).strip()
        product_code = str(row['Product Code SPMS']).strip().lower()

        # Evaluate all conditions independently
        reasons = []
        if cancel_flag == '1':
            promo_no = str(row.get('Promotion No', '')).strip()
            reasons.append(f"cancel flag is 1 (Promotion No: {promo_no})")
        if promo_check == '0':
            reasons.append("promotion number check is 0")
        if product_code == 'not found':
            reasons.append("model is 'Not Found'")
        if model_check == '0':
            reasons.append("model check is 0")

        q = row['Q']
        soa = row['SOA']
        q_times_soa = q * soa

        # Classify row and add to corresponding list and total
        if reasons:
            not_ok_rows.append(q)
            not_ok_total += q_times_soa
            comment = f"Model {model} in BEBS {bebs}: " + ", ".join(reasons)
            comments.append(comment)
        else:
            ok_rows.append(q)
            ok_total += q_times_soa

    ok_sum = sum(ok_rows)
    not_ok_sum = sum(not_ok_rows)
    final_comment = "; ".join(comments) if comments else ""

    comment_summary.append({
        'BEBS': bebs,
        'Model': model,
        'OK': ok_sum,
        'NOT OK': not_ok_sum,
        'OK TOTAL': ok_total,
        'NOT OK TOTAL': not_ok_total,
        'Comment': final_comment
    })

# Final DataFrame
models_with_comments = pd.DataFrame(comment_summary)

# Save to Excel
output_path_12 = os.path.join(folder_path, '11. Data_verified_permodel.xlsx')
models_with_comments.to_excel(output_path_12, index=False)
print(f"✅ File saved to: {output_path_12}")




























    

# Seperating into different BEBS codes
unique_BEBS = data_claim['BEBS'].unique()
username = getpass.getuser()
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # Updated timestamp format

# Function to sanitize filenames
def sanitize_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "_", str(name))

# Save each filtered BEBS dataset to a separate Excel file
for bebs_code in unique_BEBS:
    # Filter for current BEBS code
    # Filter from the verification dataset instead of original
    filtered = data_claim_verification[data_claim_verification['BEBS'] == bebs_code].copy()
    # Drop unnecessary columns before saving
    columns_to_drop = [
    'Promotion Start Week',
    'Promotion End Week',
    'Promotion Start Year',
    'Promotion End Year',
    'Promotion Start Date Formatted',
    'Promotion End Date Formatted',
    "Week's range",
    'Bill To Name Closed 12',
    'Bill To Name Closed Orders',
    'Bill To Name SPMS',
    'Cancel Flag',
    'Product Code SPMS',
    'Bill To Name Short'
    ]
    filtered.drop(columns=columns_to_drop, inplace=True, errors='ignore')


    # Skip if no data for this BEBS
    if filtered.empty:
        continue

    # Construct filename
    filename = f"{os.path.splitext(output_path)[0]}_CLAIM_{sanitize_filename(bebs_code)}_{username}_{timestamp}.xlsx"
    print(f"Saving file: {filename}")

    # Export to Excel
    filtered.to_excel(filename, index=False, sheet_name="VERIFICATION")
    print(f"File saved: {filename}")

    # ✅ Apply formatting using openpyxl
    wb = load_workbook(filename)
    ws = wb.active

    # Insert rows and columns
    ws.insert_rows(1)  # Insert row at the top
    ws.insert_cols(9)  # After column H (8th position)
    ws.insert_cols(17)  # At 17th position

    # Merge cells for section headers
    ws.merge_cells('A1:H1')
    ws.merge_cells('J1:Q1')
    ws.merge_cells('S1:Z1')

    ws['A1'] = 'Claim Data'
    ws['J1'] = 'SPMS DATA'
    ws['S1'] = 'CLOSED ORDERS + PSI DATA'

    # Header formatting
    header_font = Font(bold=True)
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = grey_fill

    # Apply white fill to all cells
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.fill = white_fill

    # Auto-adjust column widths
    for col in ws.iter_cols(min_row=2):  # Skip formatting row
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                print(f"Error processing cell {cell.coordinate}: {e}")
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save the formatted workbook
    wb.save(filename)
    print(f"✅ File formatted and saved: {filename}")

    # ✅ Add Summary per model sheet for this BEBS
    ws_summary = wb.create_sheet(title="Summary per model")

# Get summary rows for this BEBS from models_with_comments
    summary_for_bebs = models_with_comments[models_with_comments['BEBS'] == bebs_code]

# Write the summary data into the new sheet
    for r_idx, row in enumerate(dataframe_to_rows(summary_for_bebs, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_summary.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook again with the new sheet
    wb.save(filename)
    print(f"✅ Summary sheet added for BEBS {bebs_code}")


# Final confirmation message
print("All files have been saved and formatted successfully.")

